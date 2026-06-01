"""
랭킹 데이터 추출 — Excel/Google Sheets에서 뉴스레터용 데이터 가공
"""
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional
import openpyxl


@dataclass
class RankingItem:
    category: str
    name: str
    type_: str           # 니즈특화 / 기본
    recent_3m: int
    prev_3m: int
    trend: str           # ▲ 성장 / ─ 유지 / ▽ 하락
    total_17m: int
    status: str          # ✅ 노출유지 / 🟡 검토필요 / ❌ 비노출전환


@dataclass
class WeeklyDigest:
    """뉴스레터 한 호에 들어갈 데이터"""
    week_label: str                          # 예: "2026년 6월 2주차"
    top_growers: list[RankingItem] = field(default_factory=list)   # 성장률 상위 5개
    top_performers: list[RankingItem] = field(default_factory=list) # 클릭 절대값 상위 5개
    new_rankings: list[RankingItem] = field(default_factory=list)  # 신규 추가된 랭킹
    seasonal_picks: list[RankingItem] = field(default_factory=list) # 계절 추천
    stat_summary: dict = field(default_factory=dict)               # 집계 통계


def load_rankings(excel_path: str) -> list[RankingItem]:
    """nosearch_ranking_분류.xlsx 또는 마스터플랜 파일에서 전체 랭킹 로드"""
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"데이터 파일 없음: {excel_path}")

    wb = openpyxl.load_workbook(path)

    # 마스터플랜 통합본 vs 분류 파일 자동 감지
    if "🔢 572개 랭킹 전수분류" in wb.sheetnames:
        ws = wb["🔢 572개 랭킹 전수분류"]
        data_start = 7  # 헤더 포함 6행
    elif "전체 분류" in wb.sheetnames:
        ws = wb["전체 분류"]
        data_start = 2
    else:
        raise ValueError("랭킹 데이터 시트를 찾을 수 없습니다.")

    rankings = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row[2]:  # 랭킹명 없으면 스킵
            continue
        # 마스터플랜: [#, 분류, 카테고리, 랭킹명, 종류, 최근3M, 이전3M, 추세, 17M합계, ...]
        # 분류 파일:  [분류, 카테고리, 랭킹명, 종류, 최근3M, 이전3M, 추세, 17M합계, ...]
        if isinstance(row[0], int):  # 마스터플랜 (첫 컬럼이 숫자 인덱스)
            offset = 1
        else:
            offset = 0

        try:
            rankings.append(RankingItem(
                status   = str(row[0 + offset] or ""),
                category = str(row[1 + offset] or ""),
                name     = str(row[2 + offset] or ""),
                type_    = str(row[3 + offset] or ""),
                recent_3m= int(row[4 + offset] or 0),
                prev_3m  = int(row[5 + offset] or 0),
                trend    = str(row[6 + offset] or "─"),
                total_17m= int(row[7 + offset] or 0),
            ))
        except (TypeError, ValueError):
            continue

    return rankings


def build_weekly_digest(rankings: list[RankingItem], week_label: str) -> WeeklyDigest:
    """전체 랭킹 리스트에서 뉴스레터 콘텐츠용 WeeklyDigest 생성"""
    active = [r for r in rankings if "노출유지" in r.status]

    # 성장률 계산 (이전 대비)
    def growth_rate(r: RankingItem) -> float:
        if r.prev_3m == 0:
            return 0.0
        return (r.recent_3m - r.prev_3m) / r.prev_3m * 100

    # 성장 상위 5개 (최소 클릭 200 이상 필터 — 소수 이상치 제외)
    growers = sorted(
        [r for r in active if r.prev_3m >= 200],
        key=growth_rate,
        reverse=True,
    )[:5]

    # 클릭 절대값 상위 5개
    top5 = sorted(active, key=lambda r: r.recent_3m, reverse=True)[:5]

    # 계절 추천 (6~8월: 에어컨·제습기·선풍기)
    import datetime
    month = datetime.date.today().month
    seasonal_cats = _seasonal_categories(month)
    seasonal = [r for r in active if r.category in seasonal_cats]
    seasonal = sorted(seasonal, key=lambda r: r.recent_3m, reverse=True)[:5]

    stat = {
        "total_active": len(active),
        "total_rankings": len(rankings),
        "avg_recent_3m": int(sum(r.recent_3m for r in active) / len(active)) if active else 0,
        "growth_count": len([r for r in active if "성장" in r.trend]),
    }

    return WeeklyDigest(
        week_label=week_label,
        top_growers=growers,
        top_performers=top5,
        new_rankings=[],    # 신규는 별도 추적 (다음 버전)
        seasonal_picks=seasonal,
        stat_summary=stat,
    )


def _seasonal_categories(month: int) -> list[str]:
    seasonal_map = {
        (6, 7, 8): ["에어컨", "선풍기", "제습기", "에어서큘레이터"],
        (12, 1, 2): ["전기장판", "히터", "가습기", "온풍기"],
        (3, 4, 5): ["공기청정기", "로봇청소기", "청소기"],
        (9, 10, 11): ["건조기", "세탁기", "식기세척기"],
    }
    for months, cats in seasonal_map.items():
        if month in months:
            return cats
    return []


def digest_to_dict(digest: WeeklyDigest) -> dict:
    """Claude 프롬프트에 주입할 JSON 직렬화"""
    def item_dict(r: RankingItem) -> dict:
        growth = 0.0
        if r.prev_3m > 0:
            growth = round((r.recent_3m - r.prev_3m) / r.prev_3m * 100, 1)
        return {
            "category": r.category,
            "name": r.name,
            "recent_3m_clicks": r.recent_3m,
            "prev_3m_clicks": r.prev_3m,
            "growth_pct": growth,
            "trend": r.trend,
            "total_17m": r.total_17m,
        }

    return {
        "week_label": digest.week_label,
        "stats": digest.stat_summary,
        "top_growers": [item_dict(r) for r in digest.top_growers],
        "top_performers": [item_dict(r) for r in digest.top_performers],
        "seasonal_picks": [item_dict(r) for r in digest.seasonal_picks],
    }
