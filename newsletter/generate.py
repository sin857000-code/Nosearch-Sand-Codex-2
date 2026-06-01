"""
노써치 뉴스레터 패키지 생성기
실행: python generate.py

출력: 노써치_위클리_YYYYMMDD.zip
  ├── 본문_초안.txt        ← 섹션별 텍스트 (스티비 에디터에 복붙)
  ├── 편집_가이드.txt      ← 어느 섹션에 무엇을 넣을지 안내
  ├── images/
  │   ├── top5_rankings.png   (가로 막대 차트)
  │   └── growth_trend.png    (성장률 차트)
"""

import argparse
import datetime
import json
import os
import sys
import zipfile
from pathlib import Path

import anthropic
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm

# ── 한국어 폰트 설정 (없으면 기본 폰트) ──────────────────────────────────────
def setup_korean_font():
    candidates = [
        "NanumGothic", "AppleGothic", "Malgun Gothic",
        "NotoSansCJK", "Noto Sans CJK KR", "DejaVu Sans",
    ]
    available = {f.name for f in fm.fontManager.ttflist}
    for name in candidates:
        if name in available:
            plt.rcParams["font.family"] = name
            return
    plt.rcParams["axes.unicode_minus"] = False


setup_korean_font()
plt.rcParams.update({
    "figure.facecolor": "white",
    "axes.facecolor": "#F4F6F9",
    "axes.spines.top": False,
    "axes.spines.right": False,
    "axes.spines.left": False,
})

# ── 색상 팔레트 ───────────────────────────────────────────────────────────────
NAVY   = "#1B2A4A"
TEAL   = "#2E86AB"
GREEN  = "#27AE60"
RED    = "#E74C3C"
ORANGE = "#F39C12"
GRAY   = "#95A5A6"

# ── 데이터 로드 ──────────────────────────────────────────────────────────────
def load_rankings(excel_path: str) -> list[dict]:
    import openpyxl
    path = Path(excel_path)
    if not path.exists():
        print(f"❌ 파일 없음: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(path)
    if "🔢 572개 랭킹 전수분류" in wb.sheetnames:
        ws = wb["🔢 572개 랭킹 전수분류"]
        start = 7
    elif "전체 분류" in wb.sheetnames:
        ws = wb["전체 분류"]
        start = 2
    else:
        print("❌ 랭킹 시트를 찾을 수 없습니다.")
        sys.exit(1)

    rows = []
    for row in ws.iter_rows(min_row=start, values_only=True):
        if not row[2]: continue
        offset = 1 if isinstance(row[0], int) else 0
        try:
            status    = str(row[0 + offset] or "")
            category  = str(row[1 + offset] or "")
            name      = str(row[2 + offset] or "")
            recent_3m = int(row[4 + offset] or 0)
            prev_3m   = int(row[5 + offset] or 0)
            trend     = str(row[6 + offset] or "─")
            total_17m = int(row[7 + offset] or 0)
        except (TypeError, ValueError):
            continue

        growth = 0.0
        if prev_3m > 0:
            growth = round((recent_3m - prev_3m) / prev_3m * 100, 1)

        if "노출유지" in status:
            rows.append(dict(
                category=category, name=name, trend=trend,
                recent_3m=recent_3m, prev_3m=prev_3m,
                growth=growth, total_17m=total_17m,
            ))
    return rows


def get_digest(rows: list[dict], month: int) -> dict:
    seasonal_map = {
        (6,7,8): ["에어컨","선풍기","제습기"],
        (9,10,11): ["건조기","세탁기","식기세척기"],
        (12,1,2): ["전기장판","히터","가습기"],
        (3,4,5): ["공기청정기","로봇청소기"],
    }
    seasonal_cats = next(
        (cats for months, cats in seasonal_map.items() if month in months), []
    )

    top5 = sorted(rows, key=lambda r: r["recent_3m"], reverse=True)[:5]
    growers = sorted(
        [r for r in rows if r["prev_3m"] >= 200],
        key=lambda r: r["growth"], reverse=True,
    )[:5]
    seasonal = [r for r in rows if r["category"] in seasonal_cats]
    seasonal = sorted(seasonal, key=lambda r: r["recent_3m"], reverse=True)[:5]

    return dict(
        top5=top5, growers=growers, seasonal=seasonal,
        total_active=len(rows),
        growth_count=len([r for r in rows if "성장" in r["trend"]]),
        avg_click=int(sum(r["recent_3m"] for r in rows) / len(rows)) if rows else 0,
    )


# ── Claude 초안 생성 ──────────────────────────────────────────────────────────
def generate_text(digest: dict, week_label: str, api_key: str) -> dict:
    client = anthropic.Anthropic(api_key=api_key)

    grower_text = "\n".join(
        f"  {i+1}. {r['category']} — {r['name']}: {r['recent_3m']:,}클릭 ({r['growth']:+.1f}%)"
        for i, r in enumerate(digest["growers"])
    )
    top5_text = "\n".join(
        f"  {i+1}. {r['category']} — {r['name']}: {r['recent_3m']:,}클릭"
        for i, r in enumerate(digest["top5"])
    )
    seasonal_text = "\n".join(
        f"  - {r['category']} — {r['name']}: {r['recent_3m']:,}클릭"
        for r in digest["seasonal"]
    )

    prompt = f"""노써치 뉴스레터 초안을 작성해주세요.

주차: {week_label}
노출 랭킹: {digest['total_active']}개 | 성장 중: {digest['growth_count']}개 | 평균 클릭: {digest['avg_click']:,}

급상승 랭킹 TOP5:
{grower_text or '데이터 없음'}

클릭 TOP5:
{top5_text or '데이터 없음'}

계절 추천:
{seasonal_text or '해당 없음'}

아래 JSON 형식으로만 출력 (```json 없이):
{{
  "subject": "제목 (30자 이내, [노써치 위클리] 제외)",
  "opening": "인사말 (3~4문장, 트렌드 언급)",
  "highlight": "하이라이트 — 가장 주목할 랭킹 집중 소개 (4~5문장, 수치 포함)",
  "growers_intro": "급상승 랭킹 소개 도입부 (2~3문장)",
  "seasonal_copy": "계절 추천 멘트 (2~3문장)",
  "closing": "마무리 + nosearch.com 방문 유도 (2문장)"
}}"""

    msg = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1500,
        system="당신은 노써치 뉴스레터 에디터입니다. 친근하고 데이터 기반의 가전 추천 콘텐츠를 작성합니다. 이모지는 섹션당 1개 이내로 절제 사용.",
        messages=[{"role": "user", "content": prompt}],
    )

    text = msg.content[0].text.strip()
    # JSON 파싱
    if "```" in text:
        text = text.split("```")[1]
        if text.startswith("json"): text = text[4:]
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return {"subject": f"{week_label} 가전 랭킹", "opening": text,
                "highlight": "", "growers_intro": "", "seasonal_copy": "", "closing": ""}


# ── 이미지 생성 ───────────────────────────────────────────────────────────────
def make_top5_chart(top5: list[dict], out_path: Path):
    """클릭 TOP5 가로 막대 차트"""
    labels = [f"{r['category']}\n{r['name']}" for r in reversed(top5)]
    values = [r["recent_3m"] for r in reversed(top5)]
    colors = [TEAL if i < 3 else GRAY for i in range(len(top5))]

    fig, ax = plt.subplots(figsize=(8, 4))
    fig.patch.set_facecolor("white")
    bars = ax.barh(labels, values, color=list(reversed(colors)),
                   height=0.55, edgecolor="none")

    # 값 레이블
    for bar, val in zip(bars, values):
        ax.text(bar.get_width() + max(values) * 0.01, bar.get_y() + bar.get_height() / 2,
                f"{val:,}", va="center", ha="left", fontsize=10, color=NAVY, fontweight="bold")

    ax.set_xlim(0, max(values) * 1.22)
    ax.set_xlabel("최근 3개월 클릭 수", fontsize=10, color=GRAY)
    ax.set_title("📊 클릭 TOP 5 랭킹", fontsize=13, fontweight="bold",
                 color=NAVY, pad=14, loc="left")
    ax.tick_params(axis="y", labelsize=9, colors=NAVY)
    ax.tick_params(axis="x", colors=GRAY)
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{int(x):,}"))
    ax.set_facecolor("#F8FAFC")
    ax.spines["bottom"].set_color("#E0E0E0")
    ax.spines["left"].set_visible(False)

    plt.tight_layout(pad=1.2)
    plt.savefig(out_path, dpi=150, bbox_inches="tight", facecolor="white")
    plt.close()


def make_growth_chart(growers: list[dict], out_path: Path):
    """성장률 TOP5 차트"""
    if not growers:
        return

    labels = [f"{r['category']}\n{r['name']}" for r in reversed(growers)]
    values = [r["growth"] for r in reversed(growers)]
    colors = [GREEN if v >= 0 else RED for v in values]

    fig, ax = plt.subplots(figsize=(8, 4))
    fig.patch.set_facecolor("white")
    bars = ax.barh(labels, values, color=list(reversed(colors)),
                   height=0.55, edgecolor="none")

    for bar, val in zip(bars, values):
        sign = "+" if val >= 0 else ""
        ax.text(bar.get_width() + abs(max(values, key=abs)) * 0.02,
                bar.get_y() + bar.get_height() / 2,
                f"{sign}{val:.1f}%", va="center", ha="left",
                fontsize=10, color=NAVY, fontweight="bold")

    ax.axvline(0, color=GRAY, linewidth=0.8, linestyle="--")
    ax.set_xlabel("전분기 대비 성장률 (%)", fontsize=10, color=GRAY)
    ax.set_title("🚀 급상승 TOP 5 랭킹 (전분기 대비)", fontsize=13, fontweight="bold",
                 color=NAVY, pad=14, loc="left")
    ax.tick_params(axis="y", labelsize=9, colors=NAVY)
    ax.tick_params(axis="x", colors=GRAY)
    ax.set_facecolor("#F8FAFC")
    ax.spines["bottom"].set_color("#E0E0E0")
    ax.spines["left"].set_visible(False)

    max_v = max(abs(v) for v in values) if values else 100
    ax.set_xlim(-max_v * 0.1, max_v * 1.3)

    plt.tight_layout(pad=1.2)
    plt.savefig(out_path, dpi=150, bbox_inches="tight", facecolor="white")
    plt.close()


# ── 텍스트 파일 생성 ──────────────────────────────────────────────────────────
def make_draft_txt(sections: dict, digest: dict, week_label: str) -> str:
    subject = f"[노써치 위클리] {sections.get('subject', week_label + ' 가전 랭킹')}"

    top5_list = "\n".join(
        f"  {i+1}. {r['category']} — {r['name']}: {r['recent_3m']:,}클릭"
        for i, r in enumerate(digest["top5"])
    )
    growers_list = "\n".join(
        f"  {i+1}. {r['category']} — {r['name']}: {r['recent_3m']:,}클릭 ({r['growth']:+.1f}%)"
        for i, r in enumerate(digest["growers"])
    )
    seasonal_list = "\n".join(
        f"  - {r['category']} — {r['name']}: {r['recent_3m']:,}클릭"
        for r in digest["seasonal"]
    )

    return f"""━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
노써치 위클리  |  {week_label}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

【 제목 】
{subject}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【 인사말 】  ← 스티비 첫 번째 텍스트 블록
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{sections.get('opening', '')}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【 하이라이트 】  ← 스티비 두 번째 텍스트 블록
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{sections.get('highlight', '')}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【 급상승 랭킹 TOP5 】  ← 이미지(growth_trend.png) + 텍스트
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{sections.get('growers_intro', '')}

데이터:
{growers_list}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【 클릭 TOP5 】  ← 이미지(top5_rankings.png) 삽입
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
데이터:
{top5_list}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【 계절 추천 】  ← 스티비 텍스트 블록
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{sections.get('seasonal_copy', '')}

계절 추천 랭킹:
{seasonal_list}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【 마무리 】  ← 스티비 마지막 텍스트 블록
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{sections.get('closing', '')}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
이번 주 현황 요약
  노출 랭킹 수: {digest['total_active']}개
  성장 중인 랭킹: {digest['growth_count']}개
  평균 클릭 (3개월): {digest['avg_click']:,}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""


def make_guide_txt(week_label: str) -> str:
    return f"""노써치 위클리 스티비 편집 가이드  |  {week_label}
===========================================

📌 이 패키지에 포함된 파일
  - 본문_초안.txt       : 섹션별 텍스트 (복붙용)
  - images/top5_rankings.png   : 클릭 TOP5 차트 이미지
  - images/growth_trend.png    : 급상승 랭킹 차트 이미지
  - 편집_가이드.txt     : 이 파일

===========================================
📋 스티비 에디터 편집 순서
===========================================

1. app.stibee.com 로그인 → 이메일 → 새 이메일

2. 제목 입력
   → 본문_초안.txt 의 【 제목 】 복사

3. [텍스트 블록 추가] → 인사말
   → 【 인사말 】 섹션 텍스트 붙여넣기

4. [텍스트 블록 추가] → 하이라이트
   → 【 하이라이트 】 섹션 텍스트 붙여넣기

5. [이미지 블록 추가] → images/growth_trend.png 업로드
   [텍스트 블록 추가] → 【 급상승 랭킹 】 텍스트 붙여넣기

6. [이미지 블록 추가] → images/top5_rankings.png 업로드

7. [텍스트 블록 추가] → 계절 추천
   → 【 계절 추천 】 텍스트 붙여넣기

8. [버튼 블록 추가]
   텍스트: "전체 랭킹 보러가기 →"
   링크: https://nosearch.com

9. [텍스트 블록 추가] → 마무리
   → 【 마무리 】 텍스트 붙여넣기

10. 미리보기 확인 → 내용 수정 → 발송

===========================================
✏️ 수정 포인트 체크리스트
===========================================
□ 제목: 30자 이내인지 확인
□ 인사말: 계절/시기 맞는지 확인
□ 데이터 수치: 최신 데이터와 일치하는지 확인
□ 이미지: 업로드 후 모바일 미리보기 확인
□ CTA 버튼: 링크 클릭 테스트
□ 수신거부 링크: 스티비 자동 삽입 확인

===========================================
생성 일시: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}
===========================================
"""


# ── 메인 ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="노써치 뉴스레터 패키지 생성")
    parser.add_argument("--data", default="data/nosearch_data.xlsx", help="Excel 파일 경로")
    parser.add_argument("--week", default="", help="주차 레이블 (기본: 자동)")
    parser.add_argument("--out",  default=".", help="출력 폴더")
    parser.add_argument("--no-ai", action="store_true", help="Claude 없이 데이터만으로 생성")
    args = parser.parse_args()

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key and not args.no_ai:
        print("⚠️  ANTHROPIC_API_KEY 없음 — 텍스트 템플릿으로 대체합니다.")
        args.no_ai = True

    today = datetime.date.today()
    week_of_month = (today.day - 1) // 7 + 1
    week_label = args.week or f"{today.year}년 {today.month}월 {week_of_month}주차"
    date_str = today.strftime("%Y%m%d")

    # 1. 데이터 로드
    print(f"📂 데이터 로드: {args.data}")
    rows = load_rankings(args.data)
    print(f"   노출 랭킹 {len(rows)}개 로드 완료")

    digest = get_digest(rows, today.month)

    # 2. 텍스트 초안
    if args.no_ai:
        print("📝 템플릿으로 초안 생성...")
        sections = {
            "subject": f"{week_label} 급상승 랭킹 & 지금 사야 할 가전",
            "opening": f"안녕하세요! {week_label} 노써치 위클리입니다.\n이번 주도 가전 랭킹 변화와 주목할 제품을 정리했습니다.\n현재 {len(rows)}개 랭킹이 노출 중이며, {digest['growth_count']}개 랭킹이 성장세입니다.",
            "highlight": f"이번 주 가장 주목할 랭킹은 {digest['growers'][0]['category']} — {digest['growers'][0]['name']}입니다.\n전분기 대비 {digest['growers'][0]['growth']:+.1f}% 상승하며 관심이 집중되고 있습니다." if digest["growers"] else "",
            "growers_intro": "이번 주 눈에 띄게 성장한 랭킹을 소개합니다. 검색 수요 증가와 계절적 요인이 맞물려 급상승한 제품들입니다.",
            "seasonal_copy": "지금 계절에 딱 맞는 가전을 미리 비교해두세요. 성수기 품절 전에 합리적인 선택을 할 수 있습니다.",
            "closing": "이번 주도 노써치와 함께 현명한 가전 선택 하세요.\n전체 랭킹은 nosearch.com에서 확인하실 수 있습니다.",
        }
    else:
        print("🤖 Claude AI 초안 생성 중...")
        sections = generate_text(digest, week_label, api_key)
        print("   생성 완료")

    # 3. 이미지 생성
    tmp_dir = Path(args.out) / f"_tmp_{date_str}"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    img_dir = tmp_dir / "images"
    img_dir.mkdir(exist_ok=True)

    print("🖼  이미지 생성 중...")
    top5_path = img_dir / "top5_rankings.png"
    make_top5_chart(digest["top5"], top5_path)
    print(f"   top5_rankings.png")

    if digest["growers"]:
        growth_path = img_dir / "growth_trend.png"
        make_growth_chart(digest["growers"], growth_path)
        print(f"   growth_trend.png")

    # 4. 텍스트 파일 생성
    draft_path = tmp_dir / "본문_초안.txt"
    draft_path.write_text(make_draft_txt(sections, digest, week_label), encoding="utf-8")

    guide_path = tmp_dir / "편집_가이드.txt"
    guide_path.write_text(make_guide_txt(week_label), encoding="utf-8")

    # 5. ZIP 패키지 생성
    zip_name = f"노써치_위클리_{date_str}.zip"
    zip_path = Path(args.out) / zip_name

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in tmp_dir.rglob("*"):
            if f.is_file():
                zf.write(f, f.relative_to(tmp_dir))

    # 6. 임시 폴더 정리
    import shutil
    shutil.rmtree(tmp_dir)

    print(f"\n✅ 패키지 생성 완료!")
    print(f"   📦 {zip_path.resolve()}")
    print(f"\n   포함 파일:")
    with zipfile.ZipFile(zip_path) as zf:
        for name in sorted(zf.namelist()):
            info = zf.getinfo(name)
            print(f"      {name}  ({info.file_size / 1024:.1f} KB)")
    print(f"\n   👉 ZIP 압축 해제 → 편집_가이드.txt 참고하여 스티비에 업로드하세요.")


if __name__ == "__main__":
    main()
